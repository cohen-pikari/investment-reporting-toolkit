import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_formatted_excel():
    # 1. Establish the underlying operational numbers data streams
    data = {
        'Ticker': ['AAPL', 'BHP.AX', 'ALT-VC-PRV'],
        'Security Name': ['Apple Inc.', 'BHP Group Limited', 'Collins St Private Equity'],
        'Internal MV ($)': [180000.00, 212500.00, 500000.00],
        'Custodian MV ($)': [326590.00, 290058.00, 475000.00],
        'Net Variance ($)': [146590.00, 77558.00, -25000.00],
        'Exception Type': ['PRICING_BREAK', 'PRICING_BREAK', 'ALTERNATIVES_VALUATION_LAG']
    }
    
    df = pd.DataFrame(data)
    today_str = datetime.today().strftime('%Y-%m-%d')
    output_path = "output/portfolio_exception_sheet.xlsx"
    
    # 2. Initialize Workbook & Worksheet architecture layers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Exception Summary"
    ws.views.sheetView[0].showGridLines = True # Standard institutional requirement
    
    # 3. Establish Style Guidelines (Corporate Navy Theme)
    font_title = Font(name='Calibri', size=16, bold=True, color='1F497D')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11, bold=False)
    font_total = Font(name='Calibri', size=11, bold=True, color='1F497D')
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_break = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid') # Soft red highlight
    fill_lag = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Soft yellow highlight
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    double_side = Side(border_style="double", color="1F497D")
    
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=thin_side, bottom=double_side)
    
    # 4. Generate Metadata Header Layer
    ws['A1'] = "INVESTMENT OPERATIONS DAILY MIS BREAKSHEET"
    ws['A1'].font = font_title
    ws['A2'] = f"Report Date: {today_str} | Source Ledgers: IBOR vs CBOR"
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='595959')
    
    # 5. Populate Data Headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    
    # 6. Populate Operational Rows & Apply Exception Conditional Logic Matrix
    for row_idx, row_data in enumerate(df.values, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_data
            cell.border = border_cell
            
            # Context-aware alignment and formatting matrices
            if col_idx in [1, 6]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 5]:
                cell.alignment = align_right
                cell.number_format = '$#,##0.00'
            else:
                cell.alignment = align_left
                
            # Dynamic highlight row coloring depending on operational anomaly class
            if row_data[5] == 'PRICING_BREAK':
                if col_idx == 5 or col_idx == 6:
                    cell.fill = fill_break
            elif row_data[5] == 'ALTERNATIVES_VALUATION_LAG':
                if col_idx == 5 or col_idx == 6:
                    cell.fill = fill_lag

    # 7. Generate Production Total Summary Summary Block
    total_row = len(df) + 5
    ws.cell(row=total_row, column=1, value="Total Exposure").font = font_total
    ws.cell(row=total_row, column=1).alignment = align_left
    ws.cell(row=total_row, column=1).border = border_total
    
    for c in range(2, 7):
        ws.cell(row=total_row, column=c).border = border_total
        
    # Inject Excel summary mathematical formula arrays
    calc_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row-1})")
    calc_cell.font = font_total
    calc_cell.alignment = align_right
    calc_cell.number_format = '$#,##0.00'
    
    # 8. Dynamically Adjust Column Width Parameters for Visual Cleanliness
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)
    print(f"📊 S-Tier Formatted Excel Ledger successfully generated and saved to: {output_path}")

if __name__ == "__main__":
    generate_formatted_excel()

