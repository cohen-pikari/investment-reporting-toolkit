import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def run_reconciliation():
    internal_path = os.path.join('data', 'internal_ledger.csv')
    custodian_path = os.path.join('data', 'custodian_statement.csv')
    df_internal = pd.read_csv(internal_path)
    df_custodian = pd.read_csv(custodian_path)
    
    df_recon = pd.merge(df_internal, df_custodian, on='security_id', how='outer', suffixes=('_internal', '_custodian'))
    df_recon['ticker_internal'] = df_recon['ticker_internal'].fillna(df_recon['ticker_custodian'])
    df_recon['currency_internal'] = df_recon['currency_internal'].fillna(df_recon['currency_custodian'])
    
    numeric_cols = ['quantity_internal', 'market_value_internal', 'quantity_custodian', 'market_value_custodian']
    df_recon[numeric_cols] = df_recon[numeric_cols].fillna(0)
    
    df_recon['qty_variance'] = df_recon['quantity_internal'] - df_recon['quantity_custodian']
    df_recon['mv_variance'] = df_recon['market_value_internal'] - df_recon['market_value_custodian']
    
    conditions = [
        (df_recon['security_id'] == 'CASH001') & (df_recon['qty_variance'] != 0),
        (df_recon['quantity_custodian'] == 0) & (df_recon['quantity_internal'] != 0),
        (df_recon['quantity_internal'] == 0) & (df_recon['quantity_custodian'] != 0),
        (df_recon['qty_variance'] != 0),
        (df_recon['qty_variance'] == 0) & (df_recon['mv_variance'] != 0)
    ]
    choices = [
        'Cash Balance Break',
        'Unrecorded Position Break (Missing in Custodian)',
        'Unbooked Trade Break (Missing in Internal Ledger)',
        'Quantity Break (Pending/Failing Settlement)',
        'Pricing / Market Value Break'
    ]
    df_recon['exception_type'] = np.select(conditions, choices, default='Matched')
    df_exceptions = df_recon[df_recon['exception_type'] != 'Matched'].copy()
    print(f'\n[SUCCESS] Processed reconciliation. Identified {len(df_exceptions)} exceptions.')
    return df_exceptions

def export_to_excel(df_exceptions):
    output_path = os.path.join('output', 'exception_report.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Exceptions'
    ws.sheet_view.showGridLines = True
    
    font_title = Font(name='Calibri', size=16, bold=True, color='1F497D')
    font_sub = Font(name='Calibri', size=11, italic=True)
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_break = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    ws['A1'] = 'INVESTMENT OPERATIONS RECONCILIATION REPORT'
    ws['A1'].font = font_title
    ws['A2'] = 'As-Of Date: 2026-07-10 | Status: Action Required'
    ws['A2'].font = font_sub
    
    headers = ['Security ID', 'Ticker', 'Qty (Internal)', 'Qty (Custodian)', 'Qty Variance', 'MV (Internal)', 'MV (Custodian)', 'MV Variance', 'Currency', 'Exception Classification']
    ws.append([])
    ws.append(headers)
    
    for col_idx in range(1, 11):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    current_row = 5
    for _, row in df_exceptions.iterrows():
        row_data = [row['security_id'], row['ticker_internal'], row['quantity_internal'], row['quantity_custodian'], row['qty_variance'], row['market_value_internal'], row['market_value_custodian'], row['mv_variance'], row['currency_internal'], row['exception_type']]
        ws.append(row_data)
        for c_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.font = font_data
            cell.border = thin_border
            if c_idx in (3, 4, 5):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif c_idx in (6, 7, 8):
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif c_idx == 10:
                cell.fill = fill_break
                cell.font = font_bold
        current_row += 1

    col_widths = {'A': 16, 'B': 10, 'C': 15, 'D': 16, 'E': 15, 'F': 16, 'G': 16, 'H': 15, 'I': 11, 'J': 45}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    wb.save(output_path)
    print(f'[SUCCESS] Formatted report cleanly saved to: {output_path}')

if __name__ == "__main__":
    exceptions_df = run_reconciliation()
    export_to_excel(exceptions_df)
